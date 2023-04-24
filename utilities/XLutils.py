import openpyxl

#function to find the row count
def get_row_count(file,sheet):
    #load the file (Workbook)
    wb=openpyxl.load_workbook(file)
    #get the sheet from workbook
    sh=wb[sheet]
    #get the row count from the sheet
    # max_row  is keyword
    return(sh.max_row)

#function to find the column count
def get_column_count(file,sheet):
    #load the file (Workbook)
    wb=openpyxl.load_workbook(file)
    #get the sheet from workbook
    sh=wb[sheet]
    #get the column count from the sheet
    #max_column  is keyword
    return(sh.max_column)

#reading the from sheet
def read_data(file,sheet,rownum,colnum):
    #load the file (Workbook)
    wb=openpyxl.load_workbook(file)
    #get the sheet from workbook
    sh=wb[sheet]
    #retuen the data by reading from sheet
    return sh.cell(row=rownum,column=colnum).value

#writing in the sheet
def write_data(file,sheet,rownum,colnum,data):
    #load the file (Workbook)
    wb=openpyxl.load_workbook(file)
    #get the sheet from workbook
    sh=wb[sheet]
    #retuen the data by writing in the sheet
    sh.cell(row=rownum,column=colnum).value=data
    #save the file
    wb.save(file)